import os
import json
import logging
import openpyxl
import subprocess
import pandas as pd
import yfinance as yf
import investpy as inv
from datetime import datetime
from alive_progress import alive_bar
from dateutil.relativedelta import relativedelta

def log_config():
    try:
        log_directory = "logs"
        logging.basicConfig(filename='logs/logsSistem.log', encoding='utf-8', level=logging.DEBUG)
    except:
        if not os.path.exists(log_directory):
            os.makedirs(log_directory)
        log_config()
    
log_config()


inicio_processamento = datetime.now()
print("início: ", inicio_processamento)

''' função para criar estrutura de pastas, para receber dados e registrar os Logs do sistema '''
def folder_create(folder_list : list):
    try:
        if len(folder_list)>0:
            for folder in folder_list:
                if not os.path.exists(folder):
                    try:
                        os.makedirs(folder)
                    except OSError as e:
                        print(e)
                else:
                    pass
        else:
             pass 
    except Exception as e:
        logging.warning(f'main.folder_create(): {str(e)}')

''' Verificando estrutura de pastas do projeto e criando, caso não exista '''
folder_create(['logs', 'data', 'processedData'])

''' função para buscar lista de ativos da B3 '''
def get_ativos():
    try:
        brs = inv.stocks.get_stocks(country='brazil')
        lista_ativos_investpy = [] 
        for key, value in brs.iterrows():
            ativo = (value['symbol'])
            lista_ativos_investpy.append(ativo)
        cabecalho = 'Tickers'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value=cabecalho)
        
        for index, item in enumerate(lista_ativos_investpy, start=1):
            sheet.cell(row=index, column=1, value=item)
        workbook.save('data/investpy_tickers_b3.xlsx')

    except Exception as e:
        logging.warning(f'main.get_ativos(): {str(e)}')

''' Buscando ativos da B3 e gravando no arquivo -> data\tickers_investpy.xlsx'''
# get_ativos() # inserido na função reader_and_writer()

''' Definição da função que coletará os dados dos papéis da B3 '''
def obter_dados(ticker, start_date, end_date):
    try:
        # Faz o download dos dados históricos do ticker
        dados_ticker = yf.download(ticker, start=start_date, end=end_date)
        
        if dados_ticker.empty:
            logging.warning(f'Dados não encontrados para {ticker}.')
            return pd.DataFrame()  # Retorna DataFrame vazio se não houver dados

        # Adiciona a coluna do ticker e de data
        dados_ticker['Ticker'] = ticker
        dados_ticker['Data'] = dados_ticker.index

        # Reorganiza as colunas na ordem desejada
        dados_ticker = dados_ticker[['Ticker', 'Data', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']]
        
        return dados_ticker
    except Exception as e:
        logging.warning(f'Erro ao obter dados para {ticker}: {str(e)}')
        return pd.DataFrame()  # Retorna DataFrame vazio em caso de erro

def verificar_e_criar_arquivo(caminho):
    # Verifica se o arquivo existe, caso contrário, irá criar o arquivo
    try:
        open(caminho, 'a').close()  # Cria o arquivo se não existir
    except Exception as e:
        logging.warning(f'Erro ao verificar/criar arquivo: {str(e)}')

# Função principal para buscar os dados
def iniciar_busca_dados():
    try:
        caminho_tickers = 'data/investpy_tickers_b3.xlsx'

        # Carrega os tickers do arquivo Excel
        tickers_df = pd.read_excel(caminho_tickers, usecols=[0], header=None, names=['Ticker'])

        tickers = tickers_df['Ticker'].tolist()

        end_date = datetime.now()
        start_date = end_date - relativedelta(months=4)

        formated_end_date = end_date.strftime("%Y-%m-%d")
        formated_start_date = start_date.strftime("%Y-%m-%d")

        caminho_dados = 'data/yfinance_tickers_results.csv'

        # Verifica e cria arquivo para armazenar os dados
        verificar_e_criar_arquivo(caminho_dados)

        dados_totais = pd.DataFrame()

        with alive_bar(len(tickers), title="Buscando dados dos tickers") as bar:

            for ticker in tickers:

                ticker = ticker + '.SA'
                dados_ticker = obter_dados(ticker, formated_start_date, formated_end_date)
                
                if not dados_ticker.empty:
                    dados_totais = pd.concat([dados_totais, dados_ticker])
                bar()
                
        if not dados_totais.empty:
            # Salva todos os dados no arquivo CSV
            dados_totais.to_csv(caminho_dados, index=False)

    except Exception as e:
        logging.warning(f'Erro ao iniciar busca de dados: {str(e)}')

''' Executa a função principal '''
# iniciar_busca_dados() # inserido na função reader_and_writer()

''' Implementação da função para processar os dados encontrados e listar pelo maior ganho '''
def manipulador_dados():
    try:
        if os.path.exists('data/yfinance_tickers_results.csv'):
 
            df = pd.read_csv('data/yfinance_tickers_results.csv')

            unique_tickers = df['Ticker'].unique()

            sintese_dados = []

            for ticker in unique_tickers:

                sintese_ticker = []
                sintese_ticker.append(ticker)
                dados_ticker = df[df['Ticker']==ticker]

                maior_data_string = dados_ticker['Data'].max()

                maior_data = datetime.strptime(maior_data_string, '%Y-%m-%d')
                
                menor_data = dados_ticker['Data'].min()

                sintese_ticker.append(menor_data)
                dados_abertura = dados_ticker[(dados_ticker['Data']==menor_data)]
                
                valor_abertura = dados_abertura['Open']

                sintese_ticker.append(f'{((valor_abertura.values)[0]):.2f}')

                maior_data = dados_ticker['Data'].max()
                sintese_ticker.append(maior_data)
                dados_fechamento = dados_ticker[(dados_ticker['Data']==maior_data)]
                valor_fechamento = dados_fechamento['Close']
                
                sintese_ticker.append(f'{((valor_fechamento.values)[0]):.2f}')
                

                resultado_ticker = ((valor_fechamento.values)/(valor_abertura.values))
                resultado_ticker = (resultado_ticker-1)*100
                
                resultado_formatado = (f'{resultado_ticker[0]:.2f}')

                sintese_ticker.append(resultado_formatado)
                sintese_dados.append(sintese_ticker)

            colunas = ['Ticker', 'Data Inicial', 'Valor Inicial', 'Data Final', 'Valor Final', 'Variação(%)']
            dataFrame = pd.DataFrame(sintese_dados, columns=colunas)

            dataFrame['Variação(%)'] = pd.to_numeric(dataFrame['Variação(%)'])
            
            df_ordenado = dataFrame.sort_values(by='Variação(%)', ascending=False)

            df_ordenado.to_csv('processedData/tickers_ordenados.csv', index=False)
            print("finalizou manipulador de dados")
            
        else:
            print("Falha na leitura do arquivo yfinance_tickers_results.csv")

    except Exception as e:
        logging.warning(f'main.ifinanceFielExists(): {str(e)}')

# manipulador_dados() # inserido na função reader_and_writer()

def reader_and_writer():  
    data = {
        'date':''
    }
    current_date = datetime.now().strftime('%Y-%m-%d')
    
    data_folder = 'data'
    json_file = os.path.join(data_folder, 'date_search.json')
    
    if not os.path.exists(data_folder):
        os.makedirs(data_folder)
    
    if os.path.exists(json_file):
        with open(json_file, 'r') as file:
            data = json.load(file)
    # caso a data do arquivo seja menor que a data atual, executa a pesquisa e processamento dos dados
    if data['date'] < current_date:
        data_to_save = {
            'date' : current_date
        }
        with open(json_file, 'w') as file:
            json.dump(data_to_save, file, indent=4)
        get_ativos()
        iniciar_busca_dados()
        manipulador_dados()
    else: 
        print("Base de dados atualizada!")

reader_and_writer()

final_processamento = datetime.now()
tempo_processamento = final_processamento-inicio_processamento
print("\nTempo gasto no processamento: ", tempo_processamento, "\n")

arquivo = 'streamlit_arquive.py'
comando=f'streamlit run {arquivo}'
subprocess.run(comando, shell=True)

